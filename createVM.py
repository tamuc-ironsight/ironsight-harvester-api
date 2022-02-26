from IronsightHarvesterAPI import createVM
import random
import string
import sys
import openpyxl
from pprint import pprint

def listTemplates():
    print("VM Templates:\n----------------")
    for template in templates:
        print(" - " + template['normalName'] + ": " + template['description'])

if __name__ == "__main__":

    templates = []
    # Get normalName, vmName, image name, and description from excel file
    templateSheet = openpyxl.load_workbook('TemplateList.xlsx')
    # templateSheet is the first sheet (Use wb[sheetname] to get the sheet)
    templateSheet = templateSheet[templateSheet.sheetnames[0]]
    for row in range(2, templateSheet.max_row + 1):
        templates.append({
            'normalName': templateSheet.cell(row=row, column=1).value,
            'vmName': templateSheet.cell(row=row, column=2).value,
            'imageName': templateSheet.cell(row=row, column=3).value,
            'description': templateSheet.cell(row=row, column=4).value
        })

    if "--templates" in sys.argv:
        listTemplates()
        sys.exit(0)

    if len(sys.argv) < 4:
        print("Usage: python3 createVM.py [vmName] [template] [studentName]")
        sys.exit(1)
    
    vmName = sys.argv[1]
    imageName = ""
    try:
        for template in templates:
            if template['normalName'] == sys.argv[2]:
                imageName = template['imageName']
                break
        if imageName == "":
            raise Exception("Invalid template")
    except:
        print("Template not found. Here are the available templates:")
        listTemplates()
        sys.exit(1)
    userName = sys.argv[3]
    harvesterToken = "token-llcfk:dxqlvcxglvg744z2hlpvwgjfqkcxlsczprkbtg4vrxqckgdbq7z7cg"
    harvesterDomain = "https://vm.tylerharrison.dev/apis/kubevirt.io/v1/namespaces/default/virtualmachines"
    elasticDomain = "http://ssh.tylerharrison.dev:8220"
    elasticToken = "a1VJbXBYNEJZOEUtRnJyRlhvc0Q6U3VFZGRPcldUcW0tRzMyX2s3UmtTUQ=="
    vmName = vmName + "-" + userName
    randomLetters = "".join(random.choice(
    string.ascii_lowercase) for i in range(5))
    claimName = vmName + "-claim" + randomLetters
    createVM(vmName, claimName, imageName, userName, harvesterDomain, harvesterToken, elasticDomain, elasticToken)